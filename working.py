import sys

## Appending PATH of installed openpyxl Libraries
sys.path.append(r"c:\users\metap\appdata\local\programs\python\python310\lib\site-packages")

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

### NOTE: The excel file should be closed before it can be accessed

## Loading an excel notebook

wb = load_workbook('Grades.xlsx')

# '.active' acesses the active sheet in the excel sheet

ws = wb.active

## Print Cell Values

print(ws['A1'].value)
print(ws['A2'].value)

## Change a cell value and then save it

ws['A2'].value = "Test"
# Note: "ws['A2'] = 'Test' " will do the same thing
wb.save('Grades.xlsx')

## Print all sheet names in a WorkBook

print(wb.sheetnames)
sheet_names = wb.sheetnames
for item in sheet_names: print(item)

## Accessing Specific Sheets

print(wb['Sheet1'])

## Specific sheets can also be accessed as:-

ws = wb['Grades'] # 'Grade' sheet is accessed in the 'wb'
print(ws)

## Creating a sheet (Saving is a must after each modification)

#wb.create_sheet("Test")
#print(wb.sheetnames)
#wb.save('Grades.xlsx')

## Create a brand new WorkBook
wb = Workbook()
ws = wb.active # Gives the Deafult sheet when a new WorkBook is created
ws.title = "Data" # Changes the name of the active worksheet to "Data"

ws['A1'] = 'Testing' # Assigning Data like this will take a very long time

## Appending data to WorkSheet as a python list at the end of the WorkSheet
ws.append(['Pritam', 'is', 'great', '!'])
ws.append(['Python', 'is', 'great', '!'])
ws.append(['end'])

## Saving our created WorkBook
wb.save('pritam.xlsx')

## Loading a WorkBook and access a range of cells using loops

wb = load_workbook('tim.xlsx')
ws = wb.active # In actual use, load the actual WorkSheet instead of active sheet
# Specific sheets can be accessed as:- ws = wb[<sheetname>]

for row in range(1, 11): # Unlike Python, excel starts at 1.
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)]) # Print Cell info
        print(ws[char + str(row)].value) # Print Cell value
        ws[char + str(row)] = char + str(row) # Assign Cell No. as it's value 

## Saving changes to our cell

wb.save('tim_modified.xlsx')

## Merging calls 

wb = load_workbook('tim_modified.xlsx')
ws = wb.active # In actual use, load the actual WorkSheet instead of active sheet
# Specific sheets can be accessed as:- ws = wb[<sheetname>]
ws.merge_cells("A1:D1") # 'ws.unmerge_cells("A1:D1")' will unmerge cells from cell A1 to D1
# 'merge_cells' method can also merge a square range, Example: 'A1:D3'
# NOTE: Merging cells can DELETE data from some cells. Even if we unmerging it afterwards.
wb.save('tim_modified.xlsx')

## Inserting rows

wb = load_workbook('tim_modified.xlsx')
ws = wb.active # In actual use, load the actual WorkSheet instead of active sheet
# Specific sheets can be accessed as:- 'ws = wb[<sheetname>]'
ws.insert_rows(7) # Insert an empty row after row 7
ws.insert_rows(7) # Insert an empty row after row 7, Again!
# To delete row 7 use 'ws.insert_rows(7)'

wb.save('tim_modified.xlsx')

## Insert Columns

wb = load_workbook('tim_modified.xlsx')
ws = wb.active # In actual use, load the actual WorkSheet instead of active sheet
# Specific sheets can be accessed as:- ws = wb[<sheetname>]
ws.insert_cols(2) # Insert column after col B.
# NOTE: "A":1, "B":2 and so on.....
# Similarly 'ws.insert_cols(2)' will delete column B

wb.save('tim_modified.xlsx')

## Moving a range of cells

wb = load_workbook('tim_copy.xlsx')
ws = wb.active # In actual use, load the actual WorkSheet instead of active sheet
# Specific sheets can be accessed as:- ws = wb[<sheetname>]
ws.move_range("A1:D5", rows = 2, cols = 2) # Move cells in "A1:B3" to by 2 rows down
# and two columns right

wb.save("tim_copy.xlsx")

## Using Python to parse JSON data into WorkSheet

# Declaring the JSON object
data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

# Create a new empty workbook
wb = Workbook()
ws = wb.active
ws.title = "Grades"
# Take headings
headings = ['Name'] + list(data['Joe'].keys())
# Append Headings
ws.append(headings)

# Append data for each row
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# Using an excel formula to calculate and put the result in a row
for col in range(2, len(data['Joe']) + 2): # Starting at "B" to length of Data field
    char = get_column_letter(col) # Character associated with column 
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}" # Putting the SUM formula
    # In row 7, Summing from row 2 to row 6 for columns B to E

# Format Headings
for col in range(1, 6): # Again, Python starts with 0, excel starts with 1
    ws[get_column_letter(col) + '1'].font = Font(bold = True, color = "1e8f8b")
    # Colour Heading with '#1e8f8b' and make Font bold

wb.save("NewGrades.xlsx")