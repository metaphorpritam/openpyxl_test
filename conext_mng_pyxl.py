from ast import Raise
import contextlib
import sys
sys.path.append(r"C:\Users\metap\AppData\Local\Programs\Python\Python310\Lib\site-packages")

from openpyxl import Workbook, load_workbook

file_name = r"Grades.xlsx"

@contextlib.contextmanager
def open_xl_file(f_name):
    wb = load_workbook(f_name)

    yield wb

    wb.save(f_name)
    del wb

with open_xl_file(file_name) as workbook:
    sheet_names = workbook.sheetnames
    print(sheet_names)

try:
    print(workbook.sheetnames)
except:
    Raise: Exception("The variable isn't there")

    
